from __future__ import annotations

import csv
import json
import math
import re
import subprocess
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REMOTE_HOST = "c01n0006"
REMOTE_ROOT = "/share/org/SHUTUANL/shu_zhangjs/rainy model/rotating_case"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = PROJECT_ROOT / "03_inventory_tables"
OUTPUT_XLSX = OUTPUT_DIR / "rotating_case_inventory.xlsx"
OUTPUT_JSON = OUTPUT_DIR / "rotating_case_inventory_latest.json"
OUTPUT_HIGHEST_JSON = OUTPUT_DIR / "rotating_case_highest_resolution_latest.json"
OUTPUT_HIGHEST_CSV = OUTPUT_DIR / "rotating_case_highest_resolution_latest.csv"


REMOTE_CODE = r"""
import json, math, os, re
from pathlib import Path
from datetime import datetime

ROOT = Path(r'''__REMOTE_ROOT__''')

def label_to_float(s, prefix):
    if not s.startswith(prefix):
        return None
    return float(s[len(prefix):].replace('p', '.'))

def sci_label_to_float(s, prefix):
    if not s.startswith(prefix):
        return None
    return float(s[len(prefix):].replace('p', '.'))

def ffloat(s):
    return float(str(s).replace('D', 'e').replace('d', 'e'))

def read_after_header(path, header_re):
    if not path.exists():
        return ""
    lines = path.read_text(errors="ignore").splitlines()
    pat = re.compile(header_re, re.I)
    for i, line in enumerate(lines):
        if pat.search(line):
            j = i + 1
            while j < len(lines) and (not lines[j].strip() or lines[j].strip().startswith("!")):
                j += 1
            if j < len(lines):
                return lines[j].strip()
    return ""

def parse_path(run):
    rel = run.relative_to(ROOT).parts
    out = {
        "Pr_label": rel[0] if len(rel) > 0 else "",
        "Ra_label": rel[1] if len(rel) > 1 else "",
        "Ek_label": rel[2] if len(rel) > 2 else "",
        "AR_label": rel[3] if len(rel) > 3 else "",
        "Beta_label": rel[4] if len(rel) > 4 else "",
        "q_label": rel[5] if len(rel) > 5 else "",
        "grid_label": rel[6] if len(rel) > 6 else "",
        "segment_label": rel[7] if len(rel) > 8 else "main",
        "base_case_path": str(ROOT.joinpath(*rel[:7])) if len(rel) >= 7 else str(run.parent),
    }
    try:
        out["Pr"] = label_to_float(out["Pr_label"], "Pr")
        out["Ra"] = sci_label_to_float(out["Ra_label"], "Ra")
        out["Ek"] = None if out["Ek_label"] == "norotating" else sci_label_to_float(out["Ek_label"], "Ek")
        out["rotation"] = "norotating" if out["Ek"] is None else "rotating"
        out["AR"] = label_to_float(out["AR_label"], "AR")
        out["beta"] = label_to_float(out["Beta_label"], "Beta")
        qm = re.match(r"qbot(.+)_qtop(.+)", out["q_label"])
        if qm:
            out["qvapbot"] = float(qm.group(1).replace("p", "."))
            out["qvaptop"] = float(qm.group(2).replace("p", "."))
        gm = re.match(r"N(\d+)[x×](\d+)[x×](\d+)", out["grid_label"])
        if not gm:
            gm = re.match(r"N(\d+)[xX×](\d+)[xX×](\d+)", out["grid_label"])
        if gm:
            out["Nx"], out["Ny"], out["Nz"] = map(int, gm.groups())
        if out.get("Pr") and out.get("Ra") and out.get("Ek"):
            out["invRo_formula"] = math.sqrt(out["Pr"] / out["Ra"]) / out["Ek"]
            out["Ro_formula"] = 1.0 / out["invRo_formula"]
        elif out.get("Ek") is None:
            out["invRo_formula"] = 0.0
            out["Ro_formula"] = None
    except Exception as e:
        out["path_parse_warning"] = str(e)
    return out

def last_numeric_time(path):
    if not path.exists():
        return None
    last = None
    try:
        with path.open("r", errors="ignore") as f:
            for line in f:
                parts = line.split()
                if not parts:
                    continue
                try:
                    last = float(parts[0].replace("D", "e").replace("d", "e"))
                except Exception:
                    continue
    except Exception:
        return None
    return last

def parse_out_files(run):
    max_t = None
    max_ntime = None
    max_dt = None
    out_files = list(run.glob("*.out"))
    err_files = list(run.glob("*.err"))
    out_last = ""
    for f in out_files:
        text = f.read_text(errors="ignore")
        if text.strip():
            out_last = "\n".join(text.splitlines()[-8:])
        for m in re.finditer(r"T\s*=\s*([0-9.Ee+\-]+)\s+NTIME\s*=\s*(\d+)\s+DT\s*=\s*([0-9.Ee+\-]+)", text):
            t = float(m.group(1).replace("D", "e").replace("d", "e"))
            ntime = int(m.group(2))
            dt = float(m.group(3).replace("D", "e").replace("d", "e"))
            if max_t is None or t > max_t:
                max_t = t
            if max_ntime is None or ntime > max_ntime:
                max_ntime = ntime
            max_dt = dt
    err_size = sum(f.stat().st_size for f in err_files if f.exists())
    return {
        "out_file_count": len(out_files),
        "err_file_count": len(err_files),
        "err_total_bytes": err_size,
        "out_max_T": max_t,
        "out_max_NTIME": max_ntime,
        "out_last_lines": out_last,
        "out_DT": max_dt,
    }

def scan_one(run):
    row = parse_path(run)
    row["run_path"] = str(run)
    row["case_path"] = str(run.parent)
    bou = run / "bou.in"
    row["bou_exists"] = bou.exists()
    row["subjob_exists"] = (run / "subjob.sh").exists()
    row["simexec_exists"] = (run / "simexec").exists()
    grid_row = read_after_header(bou, r"\bN1\b\s+\bN2\b")
    geom_row = read_after_header(bou, r"ALX3D\s+REXT1")
    control_row = read_after_header(bou, r"\bRa\b\s+\bPr\b\s+\binvRo\b")
    boundary_row = read_after_header(bou, r"A_stopmod\s+k_stopmod\s+A_sbotmod")
    dt_row = read_after_header(bou, r"DTMAX\(dt var\.\)")
    row.update({
        "bou_grid_row": grid_row,
        "bou_geometry_row": geom_row,
        "bou_control_row": control_row,
        "bou_boundary_row": boundary_row,
        "bou_dt_row": dt_row,
    })
    try:
        vals = control_row.split()
        row["Ra_bou"] = ffloat(vals[0])
        row["Pr_bou"] = ffloat(vals[1])
        row["invRo_bou"] = ffloat(vals[2])
        row["gamma_bou"] = ffloat(vals[4])
        row["alpha_bou"] = ffloat(vals[6])
        row["beta_bou"] = ffloat(vals[7])
        row["tau_bou"] = ffloat(vals[8])
        if abs(row["invRo_bou"]) > 0.0:
            row["Ek_bou"] = math.sqrt(row["Pr_bou"] / row["Ra_bou"]) / row["invRo_bou"]
        else:
            row["Ek_bou"] = None
    except Exception as e:
        row["control_parse_warning"] = str(e)
    try:
        gvals = grid_row.split()
        row["Nx_path"] = row.get("Nx")
        row["Ny_path"] = row.get("Ny")
        row["Nz_path"] = row.get("Nz")
        row["Nx"] = int(round(ffloat(gvals[0])))
        row["Ny"] = int(round(ffloat(gvals[1])))
        row["Nz"] = int(round(ffloat(gvals[2])))
        row["grid_source"] = "bou.in"
        row["grid_points"] = row["Nx"] * row["Ny"] * row["Nz"]
    except Exception as e:
        row["grid_parse_warning"] = str(e)
    try:
        bvals = boundary_row.split()
        row["dsaltop_bou"] = ffloat(bvals[4])
        row["dsalbot_bou"] = ffloat(bvals[5])
        row["qvaptop_bou"] = ffloat(bvals[6])
        row["qvapbot_bou"] = ffloat(bvals[7])
    except Exception as e:
        row["boundary_parse_warning"] = str(e)
    try:
        row["DTMAX"] = ffloat(dt_row.split()[0])
    except Exception:
        row["DTMAX"] = None
    sj = run / "subjob.sh"
    if sj.exists():
        txt = sj.read_text(errors="ignore")
        jm = re.search(r"(?:#BSUB\s+-J|#PBS\s+-N)\s+(\S+)|#SBATCH\s+--job-name=(\S+)", txt)
        if jm:
            row["job_name"] = jm.group(1) or jm.group(2)
    row.update(parse_out_files(run))
    data = run / "data"
    time_files = {
        "avgvar_last_T": data / "avgvar.out",
        "mse_aggregation_last_T": data / "mse_aggregation.out",
        "mse_scales_last_T": data / "mse_aggregation_scales.out",
        "w_z075_length_last_T": data / "w_z075_spectral_length.out",
        "nusse_walls_last_T": data / "nusse_walls.out",
        "convective_scale_last_T": run / "diagnostics" / "scale" / "convective_scale.dat",
        "moist_integral_scale_last_T": run / "diagnostics" / "scale" / "moist_integral_scale.dat",
        "mprime_square_last_T": run / "diagnostics" / "thermo" / "mprime_square.dat",
    }
    for key, path in time_files.items():
        row[key] = last_numeric_time(path)
    times = [row.get("out_max_T")] + [row.get(k) for k in time_files]
    times = [x for x in times if isinstance(x, (int, float))]
    row["max_sim_time"] = max(times) if times else None
    row["max_NTIME"] = row.get("out_max_NTIME")
    if row["max_NTIME"] is None and row.get("max_sim_time") is not None and row.get("DTMAX"):
        row["estimated_NTIME_from_T_DT"] = int(round(row["max_sim_time"] / row["DTMAX"]))
    row["continua_file_count"] = len(list(run.glob("continua_*.h5")))
    row["non_grid_h5_count"] = len([p for p in run.glob("*.h5") if not p.name.startswith("field_grid")])
    row["has_run_output"] = bool(row.get("max_sim_time") is not None or row.get("out_max_NTIME") is not None or row.get("continua_file_count", 0) > 0)
    if row["has_run_output"]:
        row["run_status"] = "HAS_OUTPUT"
    elif row.get("out_file_count", 0) > 0:
        row["run_status"] = "SUBMITTED_NO_TIME"
    else:
        row["run_status"] = "CREATED_ONLY"
    checks = []
    try:
        if row.get("beta") is not None and row.get("dsaltop_bou") is not None:
            checks.append(abs(row["dsaltop_bou"] - (row["beta"] - 1.0)) < 1e-10)
        if row.get("dsalbot_bou") is not None:
            checks.append(abs(row["dsalbot_bou"]) < 1e-12)
        if row.get("qvaptop") is not None and row.get("qvaptop_bou") is not None:
            checks.append(abs(row["qvaptop_bou"] - row["qvaptop"]) < 1e-12)
        if row.get("qvapbot") is not None and row.get("qvapbot_bou") is not None:
            checks.append(abs(row["qvapbot_bou"] - row["qvapbot"]) < 1e-12)
        if row.get("invRo_formula") is not None and row.get("invRo_bou") is not None:
            checks.append(abs(row["invRo_bou"] - row["invRo_formula"]) < max(1e-10, abs(row["invRo_formula"]) * 1e-9))
    except Exception:
        pass
    row["parameter_check"] = "OK" if checks and all(checks) else ("INCOMPLETE" if not checks else "CHECK")
    row["scanned_at"] = datetime.now().isoformat(timespec="seconds")
    return row

rows = []
if ROOT.exists():
    for run in sorted(ROOT.rglob("run")):
        if (run / "bou.in").exists():
            try:
                rows.append(scan_one(run))
            except Exception as e:
                rows.append({"run_path": str(run), "scan_error": str(e), "scanned_at": datetime.now().isoformat(timespec="seconds")})

print(json.dumps(rows, ensure_ascii=False))
"""


def run_remote_scan() -> list[dict]:
    remote_code = REMOTE_CODE.replace("__REMOTE_ROOT__", REMOTE_ROOT)
    proc = subprocess.run(
        ["ssh", REMOTE_HOST, "python3", "-"],
        input=remote_code,
        text=True,
        capture_output=True,
        check=False,
        encoding="utf-8",
        errors="replace",
    )
    if proc.returncode != 0:
        raise RuntimeError(f"remote scan failed\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}")
    return json.loads(proc.stdout)


def sort_key(row: dict):
    ra = row.get("Ra") or 0
    ek = row.get("Ek")
    beta = row.get("beta") or 0
    ar = row.get("AR") or 0
    nx = row.get("Nx") or 0
    ek_sort = -1 if ek is None else ek
    return (ra, ek_sort, ar, beta, nx, row.get("case_path", ""))


def write_excel(rows: list[dict]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = sorted(rows, key=sort_key)
    OUTPUT_JSON.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    status_label = {
        "HAS_OUTPUT": "已跑/有输出",
        "SUBMITTED_NO_TIME": "疑似提交过但未读到时间",
        "CREATED_ONLY": "未跑/仅创建",
    }
    check_label = {
        "OK": "OK",
        "INCOMPLETE": "未完整检查",
        "CHECK": "需检查",
    }

    columns = [
        ("Ra", "Ra"),
        ("Pr", "Pr"),
        ("Ek", "Ek"),
        ("rotation", "旋转状态"),
        ("invRo_bou", "invRo(bou.in)"),
        ("Ro_formula", "Ro(公式)"),
        ("AR", "宽高比AR"),
        ("beta", "β"),
        ("qvapbot", "q_bot"),
        ("qvaptop", "q_top"),
        ("dsaltop_bou", "dsal_top"),
        ("dsalbot_bou", "dsal_bot"),
        ("Nx", "Nx"),
        ("Ny", "Ny"),
        ("Nz", "Nz"),
        ("DTMAX", "DTMAX"),
        ("has_run_output", "是否已跑"),
        ("run_status", "运行状态"),
        ("max_sim_time", "最大模拟时间T"),
        ("max_NTIME", "最大NTIME"),
        ("estimated_NTIME_from_T_DT", "估算时间步"),
        ("avgvar_last_T", "avgvar最后T"),
        ("mse_aggregation_last_T", "MSE方差最后T"),
        ("mse_scales_last_T", "MSE尺度最后T"),
        ("w_z075_length_last_T", "w尺度最后T"),
        ("nusse_walls_last_T", "Nu最后T"),
        ("out_file_count", ".out数"),
        ("continua_file_count", "continua数"),
        ("err_total_bytes", ".err字节"),
        ("parameter_check", "参数检查"),
        ("job_name", "作业名"),
        ("case_path", "case路径"),
        ("run_path", "run路径"),
        ("bou_control_row", "bou.in控制行"),
        ("bou_boundary_row", "bou.in边界行"),
        ("scanned_at", "更新时间"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "case_inventory"
    ws.append([label for _, label in columns])
    for row in rows:
        line = []
        for key, _ in columns:
            val = row.get(key)
            if key == "has_run_output":
                val = "是" if val else "否"
            elif key == "run_status":
                val = status_label.get(val, val)
            elif key == "parameter_check":
                val = check_label.get(val, val)
            line.append(val)
        ws.append(line)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(text), 80))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 55))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = wb.create_sheet("summary")
    summary["A1"] = "远端 case 清单自动更新"
    summary["A1"].font = Font(size=16, bold=True)
    summary["A3"] = "远端根目录"
    summary["B3"] = REMOTE_ROOT
    summary["A4"] = "SSH Host"
    summary["B4"] = REMOTE_HOST
    summary["A5"] = "更新时间"
    summary["B5"] = now
    summary["A6"] = "case总数"
    summary["B6"] = len(rows)
    summary["A8"] = "运行状态统计"
    summary["A8"].font = Font(bold=True)
    for i, (status, count) in enumerate(Counter(r.get("run_status", "UNKNOWN") for r in rows).items(), start=9):
        summary[f"A{i}"] = status_label.get(status, status)
        summary[f"B{i}"] = count
    start = 9 + len(Counter(r.get("run_status", "未知") for r in rows)) + 2
    summary[f"A{start}"] = "按 Ra 统计"
    summary[f"A{start}"].font = Font(bold=True)
    by_ra = defaultdict(int)
    for r in rows:
        by_ra[r.get("Ra_label", str(r.get("Ra")))] += 1
    for i, (ra, count) in enumerate(sorted(by_ra.items()), start=start + 1):
        summary[f"A{i}"] = ra
        summary[f"B{i}"] = count
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 32

    wb.save(OUTPUT_XLSX)


PROGRESS_FIELDS = [
    "out_max_T",
    "avgvar_last_T",
    "mse_aggregation_last_T",
    "mse_scales_last_T",
    "w_z075_length_last_T",
    "nusse_walls_last_T",
    "convective_scale_last_T",
    "moist_integral_scale_last_T",
    "mprime_square_last_T",
    "max_sim_time",
]


def value_key(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(f"{float(value):.12g}")
    return value


def physical_key(row: dict) -> tuple:
    return tuple(
        value_key(value)
        for value in (
            row.get("Ra_bou", row.get("Ra")),
            row.get("Pr_bou", row.get("Pr")),
            row.get("Ek_bou", row.get("Ek")),
            row.get("AR"),
            row.get("beta_bou", row.get("beta")),
            row.get("gamma_bou"),
            row.get("alpha_bou"),
            row.get("tau_bou"),
            row.get("qvapbot_bou", row.get("qvapbot")),
            row.get("qvaptop_bou", row.get("qvaptop")),
        )
    )


def resolution_key(row: dict) -> tuple:
    return physical_key(row) + (
        row.get("Nx"),
        row.get("Ny"),
        row.get("Nz"),
    )


def aggregate_resolution_rows(rows: list[dict]) -> list[dict]:
    grouped: dict[tuple, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[resolution_key(row)].append(row)

    aggregated: list[dict] = []
    for members in grouped.values():
        representative = next(
            (row for row in members if row.get("segment_label") == "main"),
            members[0],
        )
        merged = dict(representative)
        merged["segment_count"] = len(members)
        merged["segments"] = ", ".join(
            sorted({str(row.get("segment_label", "main")) for row in members})
        )
        merged["run_paths"] = "\n".join(
            sorted({str(row.get("run_path", "")) for row in members})
        )
        merged["base_case_paths"] = "\n".join(
            sorted({str(row.get("base_case_path", "")) for row in members})
        )
        merged["resolution"] = (
            f"{merged.get('Nx')}x{merged.get('Ny')}x{merged.get('Nz')}"
        )
        merged["grid_points"] = int(
            (merged.get("Nx") or 0)
            * (merged.get("Ny") or 0)
            * (merged.get("Nz") or 0)
        )
        latest_time = None
        latest_source = ""
        for field in PROGRESS_FIELDS:
            values = [
                row.get(field)
                for row in members
                if isinstance(row.get(field), (int, float))
            ]
            merged[field] = max(values) if values else None
            if values and (latest_time is None or max(values) > latest_time):
                latest_time = max(values)
                latest_source = field
        merged["max_sim_time"] = latest_time
        merged["latest_time_source"] = latest_source
        ntime_values = [
            row.get("max_NTIME")
            for row in members
            if isinstance(row.get("max_NTIME"), (int, float))
        ]
        merged["max_NTIME"] = max(ntime_values) if ntime_values else None
        merged["out_file_count"] = sum(row.get("out_file_count", 0) or 0 for row in members)
        merged["err_total_bytes"] = sum(row.get("err_total_bytes", 0) or 0 for row in members)
        merged["continua_file_count"] = sum(
            row.get("continua_file_count", 0) or 0 for row in members
        )
        merged["has_run_output"] = any(row.get("has_run_output") for row in members)
        merged["run_status"] = (
            "HAS_OUTPUT"
            if merged["has_run_output"]
            else (
                "SUBMITTED_NO_TIME"
                if any(row.get("run_status") == "SUBMITTED_NO_TIME" for row in members)
                else "CREATED_ONLY"
            )
        )
        checks = {row.get("parameter_check") for row in members}
        merged["parameter_check"] = "OK" if checks == {"OK"} else ",".join(sorted(str(x) for x in checks))
        aggregated.append(merged)
    return aggregated


def select_highest_running_resolution(rows: list[dict]) -> list[dict]:
    by_physics: dict[tuple, list[dict]] = defaultdict(list)
    for row in rows:
        if row.get("has_run_output"):
            by_physics[physical_key(row)].append(row)

    selected = []
    for candidates in by_physics.values():
        ordered = sorted(
            candidates,
            key=lambda row: (
                row.get("grid_points", 0) or 0,
                row.get("Nx", 0) or 0,
                row.get("Ny", 0) or 0,
                row.get("Nz", 0) or 0,
                row.get("max_sim_time", 0) or 0,
            ),
            reverse=True,
        )
        best = dict(ordered[0])
        best["available_resolutions"] = ", ".join(
            row.get("resolution", "") for row in ordered
        )
        best["lower_resolution_count"] = max(0, len(ordered) - 1)
        selected.append(best)
    return sorted(selected, key=sort_key)


TABLE_COLUMNS = [
    ("Ra_bou", "Ra"),
    ("Pr_bou", "Pr"),
    ("Ek_bou", "Ek"),
    ("rotation", "旋转"),
    ("AR", "AR"),
    ("beta_bou", "beta"),
    ("gamma_bou", "gamma"),
    ("alpha_bou", "alpha"),
    ("tau_bou", "tau"),
    ("qvapbot_bou", "q_bot"),
    ("qvaptop_bou", "q_top"),
    ("resolution", "网格(bou.in)"),
    ("grid_points", "总网格点"),
    ("max_sim_time", "最新物理时间T"),
    ("latest_time_source", "最新时间来源"),
    ("max_NTIME", "最大NTIME"),
    ("avgvar_last_T", "avgvar最后T"),
    ("mprime_square_last_T", "mprime最后T"),
    ("convective_scale_last_T", "lc最后T"),
    ("moist_integral_scale_last_T", "Lm最后T"),
    ("segments", "已合并段"),
    ("run_status", "状态"),
    ("parameter_check", "参数检查"),
    ("available_resolutions", "已有输出的分辨率"),
    ("base_case_paths", "远端case路径"),
]


RAW_COLUMNS = [
    ("Ra_bou", "Ra"),
    ("Pr_bou", "Pr"),
    ("Ek_bou", "Ek"),
    ("AR", "AR"),
    ("beta_bou", "beta"),
    ("qvapbot_bou", "q_bot"),
    ("qvaptop_bou", "q_top"),
    ("Nx", "Nx(bou.in)"),
    ("Ny", "Ny(bou.in)"),
    ("Nz", "Nz(bou.in)"),
    ("segment_label", "段"),
    ("max_sim_time", "最大T"),
    ("avgvar_last_T", "avgvar最后T"),
    ("mprime_square_last_T", "mprime最后T"),
    ("run_status", "状态"),
    ("parameter_check", "参数检查"),
    ("run_path", "run路径"),
    ("bou_grid_row", "bou.in网格行"),
    ("bou_control_row", "bou.in控制行"),
    ("bou_boundary_row", "bou.in边界行"),
    ("scanned_at", "扫描时间"),
]


def append_table_sheet(wb: Workbook, title: str, rows: list[dict], columns) -> None:
    ws = wb.create_sheet(title)
    ws.append([label for _, label in columns])
    for row in rows:
        ws.append([row.get(key) for key, _ in columns])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in ws[letter]
        )
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 55))
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_current_inventory(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    resolution_rows = aggregate_resolution_rows(rows)
    highest_rows = select_highest_running_resolution(resolution_rows)
    OUTPUT_JSON.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    OUTPUT_HIGHEST_JSON.write_text(
        json.dumps(highest_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    with OUTPUT_HIGHEST_CSV.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=[key for key, _ in TABLE_COLUMNS])
        writer.writeheader()
        for row in highest_rows:
            writer.writerow({key: row.get(key) for key, _ in TABLE_COLUMNS})

    wb = Workbook()
    wb.remove(wb.active)
    append_table_sheet(wb, "最高分辨率_已有输出", highest_rows, TABLE_COLUMNS)
    append_table_sheet(
        wb,
        "所有分辨率_已合并conti",
        sorted(resolution_rows, key=sort_key),
        TABLE_COLUMNS,
    )
    append_table_sheet(wb, "原始main_conti段", sorted(rows, key=sort_key), RAW_COLUMNS)
    summary = wb.create_sheet("说明", 0)
    summary["A1"] = "旋转湿对流远端算例台账"
    summary["A1"].font = Font(size=16, bold=True)
    summary["A3"] = "扫描时间"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "远端根目录"
    summary["B4"] = REMOTE_ROOT
    summary["A5"] = "原始main/conti段数"
    summary["B5"] = len(rows)
    summary["A6"] = "合并conti后的参数-分辨率组合数"
    summary["B6"] = len(resolution_rows)
    summary["A7"] = "主表case数"
    summary["B7"] = len(highest_rows)
    summary["A9"] = "主表规则"
    summary["B9"] = "仅保留已有输出的case；同一物理参数只保留bou.in中总网格点数最高的一套。"
    summary["A10"] = "进度规则"
    summary["B10"] = "合并main与所有conti，取日志、avgvar、mprime、lc、Lm等诊断中的最大物理时间。"
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 100
    wb.save(OUTPUT_XLSX)
    return highest_rows, resolution_rows


def main() -> None:
    rows = run_remote_scan()
    highest_rows, resolution_rows = write_current_inventory(rows)
    print(
        f"updated {OUTPUT_XLSX}: {len(rows)} raw segments, "
        f"{len(resolution_rows)} parameter-resolution groups, "
        f"{len(highest_rows)} highest-resolution cases with output"
    )


if __name__ == "__main__":
    main()
